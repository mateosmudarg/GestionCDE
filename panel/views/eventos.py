from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.utils.dateparse import parse_date
from django.http import HttpResponse
from django.core.serializers.json import DjangoJSONEncoder
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import date
import io, json
from ventas.models import Venta, Producto
from eventos.models import Evento

@login_required
def lista_eventos(request):
    try:
        eventos = Evento.objects.all().order_by('fecha')
        hoy = date.today()
        for evento in eventos:
            evento.es_proximo = evento.fecha >= hoy and not any(
                e.es_proximo for e in eventos if e.fecha > hoy and e != evento
            )
    except Exception:
        eventos = []
    return render(request, 'eventos/lista.html', {'eventos': eventos})


@login_required
def calendario_eventos(request):
    try:
        eventos = Evento.objects.all().select_related('gestion').order_by('fecha')
        eventos_json = [
            {
                "id": e.id,
                "title": e.nombre,
                "start": e.fecha.isoformat(),
                "description": e.descripcion,
                "lugar": e.lugar,
                "gestion": e.gestion.nombre if e.gestion else "Sin gestión",
            }
            for e in eventos
        ]
    except Exception:
        eventos = []
        eventos_json = []
    context = {
        "eventos": eventos,
        "eventos_json": json.dumps(eventos_json, cls=DjangoJSONEncoder),
    }
    return render(request, "eventos/calendario.html", context)


@login_required
def detalles_evento(request, id):
    evento = get_object_or_404(Evento, id=id)
    try:
        ventas = Venta.objects.filter(evento=evento).select_related('producto')
    except Exception:
        ventas = []

    try:
        total_ventas = ventas.count() if hasattr(ventas, 'count') else len(ventas)
        total_ganancias = sum((v.precio_unitario_venta * v.cantidad) for v in ventas)
        ganancias_netas = sum(getattr(v, 'ganancia', lambda: 0)() for v in ventas)
    except Exception:
        total_ventas = total_ganancias = ganancias_netas = 0

    # Productos vendidos
    productos_stats = {}
    for v in ventas:
        try:
            productos_stats[v.producto.nombre] = productos_stats.get(v.producto.nombre, 0) + v.cantidad
        except Exception:
            continue
    productos_labels = list(productos_stats.keys())
    productos_data = list(productos_stats.values())
    total_productos = sum(productos_data)

    # Métodos de pago
    pagos_stats = {}
    for v in ventas:
        try:
            pagos_stats[v.medio_de_pago] = pagos_stats.get(v.medio_de_pago, 0) + 1
        except Exception:
            continue
    pagos_labels = list(pagos_stats.keys())
    pagos_data = list(pagos_stats.values())

    # Datos de exportación XLSX
    productos_export = []
    for nombre, cantidad in productos_stats.items():
        try:
            precio_unitario_venta = float(next(v.precio_unitario_venta for v in ventas if v.producto.nombre == nombre))
            precio_unitario_compra = float(next(v.precio_unitario_compra for v in ventas if v.producto.nombre == nombre))
            total = float(sum(v.total() for v in ventas if v.producto.nombre == nombre))
            ganancia = float(sum(v.ganancia() for v in ventas if v.producto.nombre == nombre))
            productos_export.append({
                'Producto': nombre,
                'Cantidad': cantidad,
                'Precio Unitario Venta': precio_unitario_venta,
                'Precio Unitario Compra': precio_unitario_compra,
                'Total': total,
                'Ganancia': ganancia,
            })
        except Exception:
            continue

    pagos_export = [{'Metodo de pago': k, 'Cantidad': v} for k, v in pagos_stats.items()]

    if request.GET.get('export') == 'xlsx':
        output = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = f"Evento {evento.nombre}"

        headers = ["Producto", "Cantidad", "Precio Unitario Venta", "Precio Unitario Compra", "Total", "Ganancia"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        for item in productos_export:
            ws.append([
                item['Producto'],
                item['Cantidad'],
                item['Precio Unitario Venta'],
                item['Precio Unitario Compra'],
                item['Total'],
                item['Ganancia']
            ])

        ws.append([])
        ws.append(["", "", "", "Total bruto", float(total_ganancias)])
        ws.append(["", "", "", "Ganancia neta", float(ganancias_netas)])

        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=evento_{id}_detalle.xlsx'
        return response

    context = {
        "evento": evento,
        "total_ventas": total_ventas,
        "total_ganancias": total_ganancias,
        "ganancias_netas": ganancias_netas,
        "productos_labels": json.dumps(productos_labels, cls=DjangoJSONEncoder),
        "productos_data": json.dumps(productos_data, cls=DjangoJSONEncoder),
        "total_productos": total_productos,
        "pagos_labels": json.dumps(pagos_labels, cls=DjangoJSONEncoder),
        "pagos_data": json.dumps(pagos_data, cls=DjangoJSONEncoder),
        'productos_export': json.dumps(productos_export, cls=DjangoJSONEncoder),
        'pagos_export': json.dumps(pagos_export, cls=DjangoJSONEncoder),
    }
    return render(request, "eventos/detalles.html", context)
